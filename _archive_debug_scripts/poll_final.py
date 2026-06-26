import requests, sys, json, time, base64, io
sys.stdout.reconfigure(encoding='utf-8')

N8N_BASE = 'https://weinflow.app.n8n.cloud'
s = requests.Session()
s.post(f'{N8N_BASE}/rest/login', json={'emailOrLdapLoginId':'af8847492@gmail.com','password':'Tyzk7mra'})

def r1(pool, v):
    if isinstance(v, str) and v.isdigit():
        idx = int(v)
        if idx < len(pool): return pool[idx]
    return v

LATEST_KNOWN = 204
print('Polling for exec > 204...')
for attempt in range(40):
    try:
        r = s.get(f'{N8N_BASE}/rest/executions?workflowId=6v9BXm5uZpuJS8fd&limit=1', timeout=20)
        results = r.json().get('data', {}).get('results', [])
        if results:
            ex = results[0]
            eid = int(ex.get('id', 0))
            st = ex.get('status')
            if eid > LATEST_KNOWN:
                print(f'  [{attempt+1}] Exec {eid}: {st}')
                if st in ('success', 'error', 'crashed'):
                    print(f'\nExec {eid}: {st}')
                    r2 = s.get(f'{N8N_BASE}/rest/executions/{eid}', timeout=30)
                    full = r2.json()['data']
                    pool = json.loads(full['data'])
                    run_data = pool[5]

                    # Get Parse Offer Data output
                    ref = run_data.get('Parse Offer Data', '')
                    runs = r1(pool, ref)
                    if isinstance(runs, list) and runs:
                        ri = r1(pool, runs[0])
                        if isinstance(ri, dict) and not ri.get('error'):
                            d = r1(pool, ri.get('data', {}))
                            m = r1(pool, d.get('main', []) if isinstance(d, dict) else [])
                            p0 = r1(pool, m[0] if isinstance(m, list) and m else None)
                            i0 = r1(pool, p0[0] if isinstance(p0, list) and p0 else None)
                            j = r1(pool, i0.get('json', {}) if isinstance(i0, dict) else {})
                            od_ref = j.get('offer_data') if isinstance(j, dict) else None
                            od = r1(pool, od_ref) if isinstance(od_ref, str) else od_ref

                            if isinstance(od, dict):
                                offers_ref = od.get('offers', [])
                                offers = r1(pool, offers_ref) if isinstance(offers_ref, str) else offers_ref
                                menu_items_ref = od.get('menu_items', [])
                                menu_items = r1(pool, menu_items_ref) if isinstance(menu_items_ref, str) else menu_items_ref

                                print(f'\nOffers count: {len(offers) if isinstance(offers, list) else "?"}')

                                # Check first offer
                                if isinstance(offers, list) and offers:
                                    o0 = r1(pool, offers[0])
                                    print(f'offer[0] type: {type(o0)}')
                                    if isinstance(o0, dict):
                                        print(f'offer[0].party_size: {o0.get("party_size")}')
                                        print(f'offer[0].tier: {o0.get("tier")}')
                                        print(f'offer[0].status: {o0.get("status")}')
                                        print(f'offer[0].hook_type: {o0.get("hook_type")}')

                                print(f'menu_items: {len(menu_items) if isinstance(menu_items, list) else type(menu_items)}')

                    # Decode XLSX to confirm
                    gof_ref = run_data.get('Generate Offer Files','')
                    gof = r1(pool, gof_ref)
                    if isinstance(gof, list) and gof:
                        gi = r1(pool, gof[0])
                        if isinstance(gi, dict):
                            gd = r1(pool, gi.get('data',{}))
                            gm = r1(pool, gd.get('main',[]) if isinstance(gd,dict) else [])
                            gp0 = r1(pool, gm[0] if isinstance(gm,list) and gm else None)
                            gi0 = r1(pool, gp0[0] if isinstance(gp0,list) and gp0 else None)
                            gj = r1(pool, gi0.get('json',{}) if isinstance(gi0,dict) else {})
                            files_ref = gj.get('files',[]) if isinstance(gj,dict) else []
                            files = r1(pool, files_ref) if isinstance(files_ref,str) else files_ref
                            if isinstance(files, list):
                                print(f'\nGenerate Offer Files returned {len(files)} files')
                                for f_ref in files:
                                    f = r1(pool, f_ref)
                                    fname = r1(pool, f.get('filename','')) if isinstance(f,dict) else ''
                                    fsize = f.get('size',0) if isinstance(f,dict) else 0
                                    print(f'  {fname} ({fsize} bytes)')
                                    if isinstance(f,dict) and 'Template' in str(fname):
                                        try:
                                            import openpyxl
                                            b64 = r1(pool, f.get('content_base64',''))
                                            if isinstance(b64, str):
                                                wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(b64)))
                                                ws = wb['Offers']
                                                print(f'  Offers sheet rows 3-7:')
                                                for ri2, row in enumerate(ws.iter_rows(min_row=3, max_row=7, values_only=True)):
                                                    vals = [str(c)[:20] if c else '' for c in list(row)[1:6]]
                                                    print(f'    {vals}')
                                                # Menu Input
                                                ws2 = wb['Menu Input']
                                                data_rows = sum(1 for row in ws2.iter_rows(min_row=3, values_only=True) if any(c for c in row))
                                                print(f'  Menu Input data rows: {data_rows}')
                                        except Exception as e:
                                            print(f'  XLSX error: {e}')
                    break
            else:
                print(f'  [{attempt+1}] Still exec {eid}: {st}')
    except Exception as e:
        print(f'  [{attempt+1}] Error: {e}')
    time.sleep(10)
print('\nDone.')
