import requests, sys, json, time, base64, io
sys.stdout.reconfigure(encoding='utf-8')

N8N_BASE = 'https://weinflow.app.n8n.cloud'
s = requests.Session()
s.post(f'{N8N_BASE}/rest/login', json={'emailOrLdapLoginId':'af8847492@gmail.com','password':'Tyzk7mra'})

def resolve(pool, v, depth=0):
    if depth > 10: return v
    try:
        idx = int(v)
        if 0 <= idx < len(pool): return resolve(pool, pool[idx], depth+1)
    except: pass
    return v

LATEST = 212
print('Polling for exec > 212...')
t_start = time.time()

for attempt in range(60):
    try:
        r = s.get(f'{N8N_BASE}/rest/executions?workflowId=6v9BXm5uZpuJS8fd&limit=1', timeout=20)
        results = r.json().get('data', {}).get('results', [])
        if results:
            ex = results[0]; eid = int(ex.get('id', 0)); st = ex.get('status')
            if eid > LATEST:
                print(f'  [{attempt+1}] Exec {eid}: {st}')
                if st in ('success', 'error', 'crashed'):
                    elapsed = time.time() - t_start
                    r2 = s.get(f'{N8N_BASE}/rest/executions/{eid}', timeout=30)
                    full = r2.json()['data']

                    # Execution time from timestamps
                    started = full.get('startedAt',''); stopped = full.get('stoppedAt','')
                    print(f'\nExec {eid}: {st}')
                    print(f'Started: {started}')
                    print(f'Stopped: {stopped}')

                    pool = json.loads(full['data'])
                    run_data = resolve(pool, pool[2].get('runData'))

                    if st == 'error':
                        last = resolve(pool, pool[2].get('lastNodeExecuted', ''))
                        err = resolve(pool, pool[2].get('error', {}))
                        msg = resolve(pool, err.get('message', '')) if isinstance(err, dict) else ''
                        print(f'FAILED at: {last}')
                        print(f'Error: {repr(str(msg)[:400])}')
                        break

                    nodes_ran = len(run_data) if isinstance(run_data, dict) else 0
                    print(f'Nodes ran: {nodes_ran}')

                    # Check recall nodes ran
                    for node_name in ['5. wein-recall — Embed Query', '5b. wein-recall — Match Offers',
                                      '5c. wein-recall — Match WeIN Offers', '🔀 Merge Recall Results',
                                      '✨ 6. wein-innovate — 20 Concepts']:
                        ran = node_name in (run_data or {})
                        print(f'  {"✅" if ran else "❌"} {node_name}')

                    # Confirm advisor/intel/scout did NOT run
                    print('\nDisconnected nodes (should NOT have run):')
                    for node_name in ['🧠 3. wein-advisor — Portfolio Strategy',
                                      '🔍 4. wein-intel — Market Intelligence',
                                      '3b. wein-selector — Auto-Pick Recommendations',
                                      '5b. wein-waffarha-scout v2',
                                      '🔀 Merge Intel + Recall + Scout']:
                        ran = node_name in (run_data or {})
                        print(f'  {"❌ RAN (unexpected!)" if ran else "✅ did not run"}: {node_name}')

                    # Parse Offer Data debug fields
                    pod_ref = run_data.get('Parse Offer Data', '') if isinstance(run_data, dict) else ''
                    pod_runs = resolve(pool, pod_ref)
                    if isinstance(pod_runs, list) and pod_runs:
                        ri = resolve(pool, pod_runs[0])
                        d = resolve(pool, ri.get('data', {}) if isinstance(ri, dict) else {})
                        main = resolve(pool, d.get('main', []) if isinstance(d, dict) else [])
                        p0 = resolve(pool, main[0] if isinstance(main, list) and main else None)
                        i0 = resolve(pool, p0[0] if isinstance(p0, list) and p0 else None)
                        j = resolve(pool, i0.get('json', {}) if isinstance(i0, dict) else {})
                        if isinstance(j, dict):
                            print(f'\nParse Offer Data:')
                            for key in ['dbg_p0', 'dbg_t0', 'dbg_h0', 'dbg_cnt', 'dbg_mi']:
                                val = resolve(pool, j.get(key))
                                print(f'  {key}: {repr(val)}')

                    # XLSX check
                    gof_ref = run_data.get('Generate Offer Files', '') if isinstance(run_data, dict) else ''
                    gof = resolve(pool, gof_ref)
                    if isinstance(gof, list) and gof:
                        gi = resolve(pool, gof[0])
                        gd = resolve(pool, gi.get('data', {}) if isinstance(gi, dict) else {})
                        gm = resolve(pool, gd.get('main', []) if isinstance(gd, dict) else [])
                        gp0 = resolve(pool, gm[0] if isinstance(gm, list) and gm else None)
                        gi0 = resolve(pool, gp0[0] if isinstance(gp0, list) and gp0 else None)
                        gj = resolve(pool, gi0.get('json', {}) if isinstance(gi0, dict) else {})
                        files_ref = gj.get('files', []) if isinstance(gj, dict) else []
                        files = resolve(pool, files_ref) if isinstance(files_ref, str) else files_ref
                        if isinstance(files, list):
                            print(f'\nGenerate Offer Files: {len(files)} files')
                            for f_ref in files:
                                f = resolve(pool, f_ref)
                                if isinstance(f, dict):
                                    fname = resolve(pool, f.get('filename', ''))
                                    fsize = f.get('size', 0)
                                    print(f'  {fname} ({fsize} bytes)')
                                    if 'Template' in str(fname):
                                        try:
                                            import openpyxl
                                            b64 = resolve(pool, f.get('content_base64', ''))
                                            if isinstance(b64, str):
                                                wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(b64)))
                                                ws = wb['Offers']
                                                sizes = set(); tiers = set()
                                                for row in ws.iter_rows(min_row=3, max_row=22, values_only=True):
                                                    rv = list(row)
                                                    if rv[1]: sizes.add(rv[1])
                                                    if rv[2]: tiers.add(rv[2])
                                                print('  Offers sheet rows 3-6:')
                                                for row in ws.iter_rows(min_row=3, max_row=6, values_only=True):
                                                    vals = [str(c)[:30] if c else '' for c in list(row)[1:5]]
                                                    print(f'    {vals}')
                                                print(f'  Party sizes: {sorted(sizes)}')
                                                print(f'  Tiers: {sorted(tiers)}')
                                        except Exception as e:
                                            print(f'  XLSX error: {e}')
                    break
            else:
                print(f'  [{attempt+1}] Still exec {eid}: {st}')
    except Exception as e:
        print(f'  [{attempt+1}] Error: {e}')
    time.sleep(10)
print('\nDone.')
