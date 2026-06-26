import requests, sys, json, time, base64, io
sys.stdout.reconfigure(encoding='utf-8')

N8N_BASE = 'https://weinflow.app.n8n.cloud'
s = requests.Session()
s.post(f'{N8N_BASE}/rest/login', json={'emailOrLdapLoginId':'af8847492@gmail.com','password':'Tyzk7mra'})

def r1(pool, v):
    if isinstance(v, str) and v.isdigit():
        idx = int(v)
        if idx < len(pool): return pool[idx]
    return v

LATEST = 205
print('Polling for exec > 205...')
for attempt in range(40):
    try:
        r = s.get(f'{N8N_BASE}/rest/executions?workflowId=6v9BXm5uZpuJS8fd&limit=1', timeout=20)
        results = r.json().get('data', {}).get('results', [])
        if results:
            ex = results[0]
            eid = int(ex.get('id', 0))
            st = ex.get('status')
            if eid > LATEST:
                print(f'  [{attempt+1}] Exec {eid}: {st}')
                if st in ('success', 'error', 'crashed'):
                    print(f'\nExec {eid}: {st}')
                    r2 = s.get(f'{N8N_BASE}/rest/executions/{eid}', timeout=30)
                    full = r2.json()['data']
                    pool = json.loads(full['data'])
                    item2 = pool[2]
                    rd_ref = item2.get('runData', '5')
                    run_data = r1(pool, rd_ref)

                    if not isinstance(run_data, dict):
                        print(f'run_data not dict: {type(run_data)}')
                        break

                    if st == 'error':
                        last = r1(pool, item2.get('lastNodeExecuted',''))
                        print(f'Last node: {last}')
                        err_ref = item2.get('error','')
                        err = r1(pool, err_ref)
                        if isinstance(err, dict):
                            msg_ref = err.get('message','')
                            msg = r1(pool, msg_ref)
                            print(f'Error: {msg}')
                        break

                    # Get Parse Offer Data output
                    pod_ref = run_data.get('Parse Offer Data','')
                    pod_runs = r1(pool, pod_ref)
                    if isinstance(pod_runs, list) and pod_runs:
                        ri = r1(pool, pod_runs[0])
                        if isinstance(ri, dict):
                            d = r1(pool, ri.get('data',{}))
                            if isinstance(d, dict):
                                m = r1(pool, d.get('main',[]))
                                if isinstance(m, list) and m:
                                    p0 = r1(pool, m[0])
                                    if isinstance(p0, list) and p0:
                                        i0 = r1(pool, p0[0])
                                        if isinstance(i0, dict):
                                            j = r1(pool, i0.get('json',{}))
                                            if isinstance(j, dict):
                                                od = r1(pool, j.get('offer_data'))
                                                if isinstance(od, dict):
                                                    offers_r = od.get('offers',[])
                                                    offers = r1(pool, offers_r) if isinstance(offers_r,str) else offers_r
                                                    mi_r = od.get('menu_items',[])
                                                    menu_items = r1(pool, mi_r) if isinstance(mi_r,str) else mi_r
                                                    o0_r = (offers[0] if isinstance(offers,list) and offers else None)
                                                    o0 = r1(pool, o0_r) if isinstance(o0_r,str) else o0_r

                                                    print(f'\nParse Offer Data output:')
                                                    print(f'  offers count: {len(offers) if isinstance(offers,list) else "?"}')
                                                    print(f'  menu_items count: {len(menu_items) if isinstance(menu_items,list) else "?"}')
                                                    if isinstance(o0, dict):
                                                        print(f'  offer[0].party_size: {o0.get("party_size")}')
                                                        print(f'  offer[0].tier: {o0.get("tier")}')
                                                        print(f'  offer[0].status: {o0.get("status")}')
                                                        print(f'  offer[0].hook_type: {o0.get("hook_type")}')

                    # Check XLSX
                    gof_ref = run_data.get('Generate Offer Files','')
                    gof = r1(pool, gof_ref)
                    if isinstance(gof, list) and gof:
                        gi = r1(pool, gof[0])
                        if isinstance(gi, dict):
                            gd = r1(pool, gi.get('data',{}))
                            gm = r1(pool, gd.get('main',[]) if isinstance(gd,dict) else [])
                            gp0 = r1(pool, gm[0] if isinstance(gm,list) and gm else None)
                            gi0 = r1(pool, gp0[0] if isinstance(gp0,list) and gp0 else None)
                            gj = r1(pool, gi0.get('json',{}) if isinstance(gi0,dict) else {})
                            files_ref = gj.get('files',[]) if isinstance(gj,dict) else []
                            files = r1(pool, files_ref) if isinstance(files_ref,str) else files_ref
                            print(f'\nGenerate Offer Files: {len(files) if isinstance(files,list) else "?"} files')
                            if isinstance(files, list):
                                for f_ref in files:
                                    f = r1(pool, f_ref)
                                    if isinstance(f, dict):
                                        fname = r1(pool, f.get('filename',''))
                                        fsize = f.get('size',0)
                                        print(f'  {fname} ({fsize} bytes)')
                                        if 'Template' in str(fname):
                                            try:
                                                import openpyxl
                                                b64 = r1(pool, f.get('content_base64',''))
                                                if isinstance(b64, str):
                                                    wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(b64)))
                                                    ws = wb['Offers']
                                                    print(f'\n  XLSX Offers sheet (rows 3-7):')
                                                    for ri2, row in enumerate(ws.iter_rows(min_row=3, max_row=7, values_only=True)):
                                                        vals = [str(c)[:20] if c else '' for c in list(row)[1:6]]
                                                        print(f'    {vals}')
                                                    ws2 = wb['Menu Input']
                                                    mi_rows = sum(1 for row in ws2.iter_rows(min_row=3, values_only=True) if any(c for c in row if isinstance(c,str) and c.strip()))
                                                    print(f'  Menu Input populated rows: {mi_rows}')
                                            except Exception as e:
                                                print(f'  XLSX error: {e}')
                    break
            else:
                print(f'  [{attempt+1}] Still exec {eid}: {st}')
    except Exception as e:
        print(f'  [{attempt+1}] Request error: {e}')
    time.sleep(10)
print('\nDone.')
