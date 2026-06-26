import requests, sys, json, base64, io
sys.stdout.reconfigure(encoding='utf-8')

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.run([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
    import openpyxl

N8N_BASE = 'https://weinflow.app.n8n.cloud'
s = requests.Session()
s.post(f'{N8N_BASE}/rest/login', json={'emailOrLdapLoginId':'af8847492@gmail.com','password':'Tyzk7mra'})
r = s.get(f'{N8N_BASE}/rest/executions/204')
full = r.json()['data']
pool = json.loads(full['data'])
run_data = pool[5]

def res1(v):
    if isinstance(v, str) and v.isdigit():
        idx = int(v)
        if idx < len(pool): return pool[idx]
    return v

gof = res1(run_data.get('Generate Offer Files', ''))
run = res1(gof[0]) if isinstance(gof, list) else None
data = res1(run.get('data', {}))
main = res1(data.get('main', []))
e0 = res1(main[0]) if isinstance(main, list) else []
e00 = res1(e0[0]) if isinstance(e0, list) else {}
j = res1(e00.get('json', {}))
files_ref = res1(j.get('files', []))

# Find the template XLSX (Filled Offer Template)
for f_ref in files_ref:
    f = res1(f_ref)
    if not isinstance(f, dict): continue
    fname = res1(f.get('filename', ''))
    mime = res1(f.get('mime_type', ''))
    if 'Template' in str(fname) or 'offer_data' in str(fname).lower():
        b64_ref = f.get('content_base64', '')
        b64 = res1(b64_ref)
        if isinstance(b64, str):
            xlsx_bytes = base64.b64decode(b64)
            wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
            print(f'File: {fname}')
            print(f'Sheets: {wb.sheetnames}')
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                print(f'\n--- Sheet: {sheet_name} ({ws.max_row} rows × {ws.max_column} cols) ---')
                # Print first 5 rows
                for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                    if row_idx >= 6: break
                    print(f'  Row {row_idx+1}: {[str(c)[:30] if c is not None else "" for c in row[:10]]}')

            # Check for party_size and tier columns specifically in "Offers" sheet
            for sheet_name in wb.sheetnames:
                if 'offer' in sheet_name.lower() or 'data' in sheet_name.lower():
                    ws = wb[sheet_name]
                    headers = [cell.value for cell in ws[1]]
                    print(f'\nHeaders in {sheet_name}: {headers[:15]}')
                    if 'party_size' in headers or 'Party Size' in headers:
                        ps_col = headers.index('party_size') if 'party_size' in headers else headers.index('Party Size')
                        tier_col = headers.index('tier') if 'tier' in headers else (headers.index('Tier') if 'Tier' in headers else -1)
                        print('party_size values:')
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            if row[ps_col]:
                                print(f'  {row[ps_col]} | {row[tier_col] if tier_col >= 0 else "?"}')
        break
