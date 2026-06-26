import openpyxl
wb = openpyxl.load_workbook("Soho House Cairo - Filled Offer Template - Claude v1.xlsx")
print("Sheets:", wb.sheetnames)

print("\n--- Menu Input (Sheet 2) ---")
ws2 = wb["Menu Input"]
for row in ws2.iter_rows(min_row=2, max_row=6, min_col=2, max_col=9, values_only=True):
    print(row)

print("\n--- Offers (Sheet 3) headers + offer 1 row ---")
ws3 = wb["Offers"]
for row in ws3.iter_rows(min_row=2, max_row=2, min_col=2, max_col=16, values_only=True):
    print("HEADERS:", row)
for row in ws3.iter_rows(min_row=3, max_row=3, min_col=2, max_col=16, values_only=True):
    print("OFFER1: ", row)

print("\n--- Provider xlsx ---")
wb2 = openpyxl.load_workbook("Soho House Cairo - Provider Negotiation Offers - Claude v1.xlsx")
ws = wb2["Provider Offers"]
for row in ws.iter_rows(min_row=5, max_row=6, min_col=1, max_col=10, values_only=True):
    print(row)
