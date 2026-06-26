import requests, sys, json, time, base64, io
sys.stdout.reconfigure(encoding='utf-8')

N8N_BASE = 'https://weinflow.app.n8n.cloud'
s = requests.Session()
s.post(f'{N8N_BASE}/rest/login', json={'emailOrLdapLoginId':'af8847492@gmail.com','password':'Tyzk7mra'})

def resolve(pool, v, depth=0):
    if depth > 10: return v
    if isinstance(v, str) and v.isdigit():
        idx = int(v)
        if idx < len(pool):
            return resolve(pool, pool[idx], depth+1)
    return v

LATEST = 208
print('Polling for exec > 208...')
for attempt in range(50):
    try:
        r = s.get(f'{N8N_BASE}/rest/executions?workflowId=6v9BXm5uZpuJS8fd&limit=1', timeout=20)
        results = r.json().get('data', {}).get('results', [])
        if results:
            ex = results[0]
            eid = int(ex.get('id', 0))
            st = ex.get('status')
            if eid > LATEST:
                print(f'  [{attempt+1}] Exec {eid}: {st}')
                if st in ('success', 'error', 'crashed'):
                    r2 = s.get(f'{N8N_BASE}/rest/executions/{eid}', timeout=30)
                    full = r2.json()['data']

                    # Check snapshot
                    snap_nodes = full.get('workflowData', {}).get('nodes', [])
                    for n in snap_nodes:
                        if n.get('name') == 'Parse Offer Data':
                            code = n['parameters'].get('jsCode', '')
                            print(f'Snapshot code length: {len(code)}, has dbg_p0: {"dbg_p0" in code}')
                            break

                    pool = json.loads(full['data'])
                    run_data = resolve(pool, pool[2].get('runData'))

                    if st == 'error':
                        item2 = pool[2]
                        last = resolve(pool, item2.get('lastNodeExecuted', ''))
                        print(f'Failed at: {last}')
                        err = resolve(pool, item2.get('error', {}))
                        if isinstance(err, dict):
                            msg = resolve(pool, err.get('message', ''))
                            print(f'Error: {msg}')
                        break

                    # Parse Offer Data
                    pod_ref = run_data.get('Parse Offer Data', '')
                    pod_runs = resolve(pool, pod_ref)
                    if isinstance(pod_runs, list) and pod_runs:
                        ri = resolve(pool, pod_runs[0])
                        d = resolve(pool, ri.get('data', {}) if isinstance(ri, dict) else {})
                        main = resolve(pool, d.get('main', []) if isinstance(d, dict) else [])
                        p0 = resolve(pool, main[0] if isinstance(main, list) and main else None)
                        i0 = resolve(pool, p0[0] if isinstance(p0, list) and p0 else None)
                        j = resolve(pool, i0.get('json', {}) if isinstance(i0, dict) else {})

                        if isinstance(j, dict):
                            print(f'\nj keys: {list(j.keys())}')
                            for key in ['dbg_p0', 'dbg_t0', 'dbg_h0', 'dbg_cnt', 'dbg_mi']:
                                val = resolve(pool, j.get(key))
                                print(f'  {key}: {repr(val)}')

                    # XLSX
                    gof_ref = run_data.get('Generate Offer Files', '')
                    gof = resolve(pool, gof_ref)
                    if isinstance(gof, list) and gof:
                        gi = resolve(pool, gof[0])
                        gd = resolve(pool, gi.get('data', {}) if isinstance(gi, dict) else {})
                        gm = resolve(pool, gd.get('main', []) if isinstance(gd, dict) else [])
                        gp0 = resolve(pool, gm[0] if isinstance(gm, list) and gm else None)
                        gi0 = resolve(pool, gp0[0] if isinstance(gp0, list) and gp0 else None)
                        gj = resolve(pool, gi0.get('json', {}) if isinstance(gi0, dict) else {})
                        files_ref = gj.get('files', []) if isinstance(gj, dict) else []
                        files = resolve(pool, files_ref) if isinstance(files_ref, str) else files_ref
                        if isinstance(files, list):
                            print(f'\nGenerate Offer Files: {len(files)} files')
                            for f_ref in files:
                                f = resolve(pool, f_ref)
                                if isinstance(f, dict):
                                    fname = resolve(pool, f.get('filename', ''))
                                    fsize = f.get('size', 0)
                                    print(f'  {fname} ({fsize} bytes)')
                                    if 'Template' in str(fname):
                                        try:
                                            import openpyxl
                                            b64 = resolve(pool, f.get('content_base64', ''))
                                            if isinstance(b64, str):
                                                wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(b64)))
                                                ws = wb['Offers']
                                                print('  Offers sheet rows 3-7:')
                                                for row in ws.iter_rows(min_row=3, max_row=7, values_only=True):
                                                    vals = [str(c)[:30] if c else '' for c in list(row)[1:6]]
                                                    print(f'    {vals}')
                                                ws2 = wb['Menu Input']
                                                mi_rows = sum(1 for row in ws2.iter_rows(min_row=3, values_only=True) if any(c for c in row if isinstance(c, str) and c.strip()))
                                                print(f'  Menu Input rows: {mi_rows}')
                                        except Exception as e:
                                            print(f'  XLSX error: {e}')
                    break
            else:
                print(f'  [{attempt+1}] Still exec {eid}: {st}')
    except Exception as e:
        print(f'  [{attempt+1}] Error: {e}')
    time.sleep(10)
print('\nDone.')
