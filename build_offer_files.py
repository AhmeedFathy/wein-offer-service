"""
WeIN Offer File Builder
=======================
Generates all 4 output files from offer data.
Usage: python build_offer_files.py <provider_name> <vertical> <output_dir>

Reads offer data from offer_data.json in the output_dir.
Copies master templates and fills all sheets correctly.
"""

import openpyxl
import openpyxl.styles as styles
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import json
import shutil
import os
import sys
from pathlib import Path
from datetime import datetime

# ─── PATHS ───────────────────────────────────────────────────────────────────
TEMPLATES_DIR = Path(__file__).resolve().parent
TEMPLATE_FILLED   = TEMPLATES_DIR / "Template - Filled Offer Template.xlsx"
TEMPLATE_PROVIDER = TEMPLATES_DIR / "Template - Provider Negotiation Offers.xlsx"

# ─── COLORS ──────────────────────────────────────────────────────────────────
NAVY   = "1F3864"
BLUE   = "2E75B6"
TEAL   = "1A7A6E"
ORANGE = "F39C12"
WHITE  = "FFFFFF"
LGREY  = "F2F3F4"
LGREEN = "D5F5E3"
YELLOW = "FFF9C4"

# ── Schema-neutral field helpers ─────────────────────────────────────────────
def _alc(o):
    """Regular / à la carte price — handles both field name conventions."""
    return o.get("regular_egp") or o.get("price_original_egp", 0)

def _promo(o):
    """Promo price — handles both field name conventions."""
    return o.get("promo_egp") or o.get("price_discounted_egp", 0)

def _disc(o):
    """Discount as a decimal (0.28) — normalises integer storage (28) too."""
    d = o.get("discount_pct", 0)
    return d / 100 if d > 1 else d

def _disc_pct_str(o):
    """Formatted discount string e.g. '28%'."""
    return f"{_disc(o):.0%}"

def _waf_ref(o):
    pct = round(_disc(o) * 100)
    return f"{max(0, pct - 6)}%"

def _gap_str(o):
    pct = round(_disc(o) * 100)
    base = max(0, pct - 6)
    return f"+{pct - base}% ✓"

def _contents(o):
    """Bundle contents — derived from items[] (current node-8 schema), falling
    back to a legacy 'contents' string if present."""
    items = o.get("items")
    if items:
        return ", ".join(item.get("name", "") for item in items)
    return o.get("contents", "")

def _hook_line(o):
    """Guest-facing marketing hook — current schema stores this as 'hook'."""
    return o.get("hook") or o.get("hook_line", "")

def _hook_type(o):
    """Hook type/category — current schema has no hook_type; title_strategy is
    the closest available classification."""
    return o.get("hook_type") or o.get("title_strategy", "")

def _party_size(o):
    """Party size — current schema has no party_size field; derive from
    keywords in the title/hook, defaulting to '—' if unclear."""
    if o.get("party_size"):
        return o["party_size"]
    text = (o.get("title", "") + " " + o.get("hook", "")).lower()
    if "family" in text:
        return "Family"
    if "group" in text:
        return "Group"
    if "couple" in text:
        return "Couple"
    if "solo" in text:
        return "Solo"
    return "—"

def _flatten_menu_items(offers):
    """Derive a Menu Input list by flattening offers[].items[], deduplicated
    by item name (current schema has no top-level menu_items)."""
    all_items = {}
    for o in offers:
        for item in o.get("items", []):
            name = item.get("name", "")
            if name and name not in all_items:
                all_items[name] = {
                    "name": name,
                    "me_class": item.get("me_class", ""),
                    "bundle_role": item.get("bundle_role", ""),
                }
    return list(all_items.values())
# ─────────────────────────────────────────────────────────────────────────────

def navy_font(bold=True):  return Font(color=WHITE, bold=bold)
def blue_font(bold=True):  return Font(color=BLUE, bold=bold)
def teal_font(bold=True):  return Font(color=TEAL, bold=bold)
def navy_fill():           return PatternFill("solid", fgColor=NAVY)
def blue_fill():           return PatternFill("solid", fgColor=BLUE)
def teal_fill():           return PatternFill("solid", fgColor=TEAL)
def lgrey_fill():          return PatternFill("solid", fgColor=LGREY)
def yellow_fill():         return PatternFill("solid", fgColor=YELLOW)
def lgreen_fill():         return PatternFill("solid", fgColor=LGREEN)
def wrap():                return Alignment(wrap_text=True, vertical="top")
def center():              return Alignment(horizontal="center", vertical="center")
def thin_border():
    s = Side(style="thin", color="D5D8DC")
    return Border(left=s, right=s, top=s, bottom=s)

def next_version(output_dir, provider_name, file_type, ext):
    """Find next version number for output file."""
    v = 1
    while (output_dir / f"{provider_name} - {file_type} - Claude v{v}.{ext}").exists():
        v += 1
    return v

def build_filled_template(wb_out, data):
    """Build the 4-sheet Filled Offer Template."""
    provider = data["provider"]
    vertical = data["vertical"]
    location = data.get("location", "Sharm El Sheikh, Egypt")
    eur_rate = data.get("eur_rate", 62.44)
    commission = data.get("commission", 0.15)
    waffarha_adj = data.get("waffarha_adj", "+5 to +7 pts above baseline")
    offers = data.get("offers", [])
    menu_items = data.get("menu_items") or _flatten_menu_items(offers)
    # Filled template shows ALL offers (up to 20 selected + 2 backups)
    selected = [o for o in offers if o.get("status", "Selected") == "Selected"]
    backups  = [o for o in offers if o.get("status", "Selected") == "Backup"]

    # MINIMUM 13 OFFERS RULE — if under 13 total, flag it
    total_offers = len(selected) + len(backups)
    if total_offers < 13:
        print(f"  WARNING: Only {total_offers} offers in filled template — minimum is 13. Add more offers to offer_data.json.")

    # No cap on internal file — show everything

    # Remove default sheet if exists
    for sname in ["Sheet", "Sheet1"]:
        if sname in wb_out.sheetnames:
            del wb_out[sname]

    # ── SHEET 1: OVERVIEW ────────────────────────────────────────────────────
    ws = wb_out.create_sheet("Overview")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20

    # Title
    ws.merge_cells("B1:E1")
    ws["B1"] = f"{provider.upper()} — Offer Engineering Overview"
    ws["B1"].font = Font(bold=True, size=14, color=NAVY)
    ws["B1"].fill = lgrey_fill()
    ws["B1"].alignment = center()
    ws.row_dimensions[1].height = 28

    # Provider info
    info_rows = [
        ("Provider Name", provider),
        ("Category", data.get("category", vertical)),
        ("Location", location),
        ("Service Type", data.get("service_type", vertical)),
        ("Tourist Split", data.get("tourist_split", "~90% international tourists")),
        ("Certifications", data.get("certifications", "")),
        ("Key Rule", data.get("key_rule", "")),
        ("Commission Rate", f"{int(commission*100)}%"),
        ("EUR Rate", f"1 EUR = {eur_rate} EGP"),
        ("Waffarha Adj", waffarha_adj),
        ("Party Sizes", data.get("party_sizes", "Solo / Couple / Group")),
    ]
    for i, (label, value) in enumerate(info_rows, start=2):
        ws[f"B{i}"] = label
        ws[f"B{i}"].font = Font(bold=True, color=NAVY)
        ws[f"C{i}"] = value
        ws[f"C{i}"].font = Font(color="333333")

    # KPI block
    row = len(info_rows) + 3
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"] = "OFFER SET KPIs"
    ws[f"B{row}"].font = Font(bold=True, color=WHITE)
    ws[f"B{row}"].fill = navy_fill()
    ws[f"B{row}"].alignment = center()
    ws.row_dimensions[row].height = 20

    row += 1
    kpi_headers = ["Total Offers", "Avg Discount", "Highest Ticket (ALC)", "Platform Revenue (avg/order)", "Waffarha Gap Range"]
    for j, h in enumerate(kpi_headers):
        cell = ws.cell(row=row, column=j+2)
        cell.value = h
        cell.font = Font(bold=True, color=NAVY, size=9)
        cell.fill = lgrey_fill()
        cell.alignment = center()
        cell.border = thin_border()

    row += 1
    avg_disc = sum(_disc(o) for o in selected) / len(selected) if selected else 0
    highest  = max((_alc(o) for o in selected), default=0)
    avg_rev  = sum(_promo(o) * commission for o in selected) / len(selected) if selected else 0
    kpi_vals = [
        f"{len(selected)} selected + {len(backups)} backups",
        f"{avg_disc:.0%}",
        f"EGP {highest:,.0f}",
        f"EGP {avg_rev:,.0f} avg",
        waffarha_adj,
    ]
    for j, v in enumerate(kpi_vals):
        cell = ws.cell(row=row, column=j+2)
        cell.value = v
        cell.font = Font(bold=True, size=11, color=NAVY)
        cell.alignment = center()
        cell.border = thin_border()
    ws.row_dimensions[row].height = 24

    # Science hooks reference
    row += 2
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"] = "SCIENCE HOOKS REFERENCE"
    ws[f"B{row}"].font = Font(bold=True, color=WHITE)
    ws[f"B{row}"].fill = teal_fill()
    ws[f"B{row}"].alignment = center()
    ws.row_dimensions[row].height = 20

    row += 1
    for hdr, col in [("Hook", "B"), ("Source", "C"), ("Apply To", "D"), ("Rule", "E")]:
        ws[f"{col}{row}"] = hdr
        ws[f"{col}{row}"].font = Font(bold=True, color=TEAL)
        ws[f"{col}{row}"].fill = lgrey_fill()
        ws[f"{col}{row}"].border = thin_border()

    hooks = [
        ("Zero-Price Effect",  "Ariely 2007",          "Solo Entry / School Entry",   "Free add-on triggers disproportionate response — free beats % off"),
        ("Anchor Pricing",     "Tversky & Kahneman",   "All Core",                    "Show à la carte total first — promo looks massive by contrast"),
        ("Loss Aversion",      "Kahneman 1979",         "Solo/Couple Premium",         "Lead with Save X EGP — on high-ticket items this is enormous"),
        ("Experience Frame",   "Pine & Gilmore 1998",   "Couple Entry",                "Name it as a shared milestone, not a discount"),
        ("Decoy Effect",       "Huber 1982",            "Couple Core",                 "Must beat 2× solo price — makes solo look wasteful"),
        ("Reciprocity",        "Cialdini 1984",         "Couple Premium",              "Premium item framed as provider's gift for committing"),
        ("Per-Person Anchor",  "Wansink 1996",          "Group Entry",                 "Show per-person EUR price — large totals feel manageable"),
        ("Compromise Effect",  "Simonson 1989",         "All Core",                    "Middle of 3 always wins — design entry & premium around the core"),
        ("Sharing Utility",    "—",                     "Group Premium",               "One group, one journey — shared identity drives commitment"),
        ("Mental Accounting",  "Thaler 1985",           "Facility Core",               "One price covers all — removes mental calculation pain"),
        ("Host Pride",         "—",                     "Facility Premium",            "Speak to aspiration and identity, never discount language"),
    ]
    for hook_row in hooks:
        row += 1
        for j, val in enumerate(hook_row):
            cell = ws.cell(row=row, column=j+2)
            cell.value = val
            cell.font = Font(color="333333", size=9)
            cell.alignment = wrap()
            cell.border = thin_border()
        ws.row_dimensions[row].height = 28

    # ── SHEET 2: MENU INPUT ──────────────────────────────────────────────────
    ws2 = wb_out.create_sheet("Menu Input")
    ws2.column_dimensions["A"].width = 2
    ws2.column_dimensions["B"].width = 35
    ws2.column_dimensions["C"].width = 22
    ws2.column_dimensions["D"].width = 14
    ws2.column_dimensions["E"].width = 14
    ws2.column_dimensions["F"].width = 12
    ws2.column_dimensions["G"].width = 16
    ws2.column_dimensions["H"].width = 12
    ws2.column_dimensions["I"].width = 35

    ws2.merge_cells("B1:I1")
    ws2["B1"] = "MENU INPUT — Service Catalogue & Classification"
    ws2["B1"].font = Font(bold=True, size=13, color=WHITE)
    ws2["B1"].fill = navy_fill()
    ws2["B1"].alignment = center()
    ws2.row_dimensions[1].height = 26

    headers2 = ["Service Name", "Category", "Price (EGP)", "Cost Sensitivity", "ME Class", "Bundle Role", "Eligible?", "Notes"]
    for j, h in enumerate(headers2):
        cell = ws2.cell(row=2, column=j+2)
        cell.value = h
        cell.font = Font(bold=True, color=NAVY)
        cell.fill = lgrey_fill()
        cell.alignment = center()
        cell.border = thin_border()
    ws2.row_dimensions[2].height = 18

    for i, item in enumerate(menu_items, start=3):
        me_class = item.get("me_class", "")
        fill = lgrey_fill() if i % 2 == 0 else PatternFill("solid", fgColor="FAFAFA")

        # Derive cost_sensitivity from price if not already set
        raw_price = item.get("price", item.get("price_egp"))
        cs = item.get("cost_sensitivity") or ""
        if not cs or cs == "N/A":
            try:
                p = float(str(raw_price).replace(",", ""))
                if p < 100:   cs = "Low"
                elif p < 250: cs = "Low-Medium"
                elif p < 400: cs = "Medium"
                elif p < 600: cs = "High"
                else:         cs = "Very High"
            except (TypeError, ValueError):
                cs = "N/A"

        vals = [
            item.get("name", ""),
            item.get("category", "N/A"),
            raw_price if raw_price is not None else "N/A",
            cs,
            me_class,
            item.get("bundle_role", ""),
            item.get("eligible", "N/A"),
            item.get("notes", ""),
        ]
        for j, v in enumerate(vals):
            cell = ws2.cell(row=i, column=j+2)
            cell.value = v
            cell.alignment = wrap()
            cell.border = thin_border()
            cell.fill = fill
        ws2.row_dimensions[i].height = 16

    # ME Legend
    leg_row = len(menu_items) + 4
    ws2.merge_cells(f"B{leg_row}:I{leg_row}")
    ws2[f"B{leg_row}"] = "ME CLASS LEGEND"
    ws2[f"B{leg_row}"].font = Font(bold=True, color=WHITE)
    ws2[f"B{leg_row}"].fill = navy_fill()
    ws2[f"B{leg_row}"].alignment = center()

    for i, (cls, desc) in enumerate([
        ("⭐ Star",      "High popularity + High margin — Hero items — Anchor the bundle, discount 25–35%"),
        ("🐴 Plowhorse", "High popularity + Low margin — Support/Filler — Include at face value, never hero"),
        ("🧩 Puzzle",    "Low popularity + High margin — Surface in premium — Reciprocity gift"),
        ("🐕 Dog",       "Low popularity + Low margin — Never bundle — Paid upgrade only"),
    ], start=leg_row+1):
        ws2[f"B{i}"] = cls
        ws2[f"B{i}"].font = Font(bold=True, color=NAVY)
        ws2.merge_cells(f"C{i}:I{i}")
        ws2[f"C{i}"] = desc
        ws2[f"C{i}"].font = Font(color="444444", size=9)
        ws2.row_dimensions[i].height = 16

    # ── SHEET 3: OFFERS ──────────────────────────────────────────────────────
    ws3 = wb_out.create_sheet("Offers")
    ws3.merge_cells("B1:R1")
    ws3["B1"] = f"OFFERS ENGINE — {len(selected)} Selected + {len(backups)} Backups = {len(offers)} Total | INTERNAL USE ONLY — Provider sees selected offers only"
    ws3["B1"].font = Font(bold=True, size=12, color=WHITE)
    ws3["B1"].fill = navy_fill()
    ws3["B1"].alignment = center()
    ws3.row_dimensions[1].height = 24

    off_headers = ["#", "Party Size", "Offer Title", "Offer Contents", "Discount %",
                   "À la Carte EGP", "Promo Price EGP", "Save EGP", "Your 15%",
                   "Provider Gets", "Waffarha Ref %", "Gap vs Waffarha", "Hook Type",
                   "Guest Hook Line", "Validity / Terms"]
    col_widths = [5, 10, 35, 50, 10, 14, 14, 12, 12, 14, 14, 14, 18, 40, 40]
    for j, (h, w) in enumerate(zip(off_headers, col_widths)):
        col_letter = chr(66 + j)  # B=66
        ws3.column_dimensions[col_letter].width = w
        cell = ws3.cell(row=2, column=j+2)
        cell.value = h
        cell.font = Font(bold=True, color=WHITE, size=9)
        cell.fill = blue_fill()
        cell.alignment = center()
        cell.border = thin_border()
    ws3.row_dimensions[2].height = 20

    all_offers_ordered = selected + backups
    total_alc = 0
    total_promo = 0
    total_save = 0
    total_comm = 0

    for i, o in enumerate(all_offers_ordered, start=3):
        num = o.get("id", i-2)
        if o.get("status") == "Backup":
            num = f"B{i - len(selected) - 2}"
        disc    = _disc(o)
        alc     = _alc(o)
        promo   = _promo(o)
        save    = alc - promo
        comm    = round(promo * commission, 2)
        prov    = promo - comm
        waf_ref = _waf_ref(o)
        gap     = _gap_str(o)

        if o.get("status") != "Backup":
            total_alc   += alc
            total_promo += promo
            total_save  += save
            total_comm  += comm

        fill = lgrey_fill() if i % 2 == 0 else PatternFill("solid", fgColor="FAFAFA")
        if o.get("status") == "Backup":
            fill = PatternFill("solid", fgColor="FFF9C4")

        vals = [num, _party_size(o), o.get("title",""),
                _contents(o), _disc_pct_str(o), alc, promo, save,
                comm, prov, waf_ref, gap, _hook_type(o),
                _hook_line(o), o.get("terms","")]

        for j, v in enumerate(vals):
            cell = ws3.cell(row=i, column=j+2)
            if isinstance(v, list):
                cell.value = ', '.join(str(item) for item in v)
            else:
                cell.value = v
            cell.font = Font(size=9)
            cell.alignment = wrap()
            cell.border = thin_border()
            cell.fill = fill
        ws3.row_dimensions[i].height = 45

    # Totals row
    tot_row = len(all_offers_ordered) + 3
    ws3.merge_cells(f"B{tot_row}:E{tot_row}")
    ws3[f"B{tot_row}"] = f"TOTALS (selected {len(selected)}):"
    ws3[f"B{tot_row}"].font = Font(bold=True, color=NAVY)
    ws3[f"G{tot_row}"] = f"Avg {avg_disc:.1%}"
    ws3[f"H{tot_row}"] = total_alc
    ws3[f"I{tot_row}"] = total_promo
    ws3[f"J{tot_row}"] = total_save
    ws3[f"K{tot_row}"] = round(total_comm, 2)
    for col in "GHIJK":
        ws3[f"{col}{tot_row}"].font = Font(bold=True, color=NAVY)
        ws3[f"{col}{tot_row}"].border = thin_border()

    # Editable note
    ws3[f"G{tot_row+1}"] = "⬆ ONLY THIS COLUMN IS EDITABLE"
    ws3[f"G{tot_row+1}"].font = Font(italic=True, color="888888", size=8)

    # ── SHEET 4: RECOMMENDED PICKS ───────────────────────────────────────────
    ws4 = wb_out.create_sheet("Recommended Picks")
    ws4.merge_cells("B1:K1")
    ws4["B1"] = f"RECOMMENDED PICKS — {len(selected)} Selected + {len(backups)} Internal Backups"
    ws4["B1"].font = Font(bold=True, size=13, color=WHITE)
    ws4["B1"].fill = navy_fill()
    ws4["B1"].alignment = center()
    ws4.row_dimensions[1].height = 26

    ws4.merge_cells("B2:K2")
    ws4["B2"] = f"Top {len(selected)} offers for maximum conversion + commission. Sorted by promo price ascending."
    ws4["B2"].font = Font(italic=True, color=TEAL, size=9)
    ws4["B2"].alignment = center()

    pick_headers = ["#", "Offer Title", "Party Size", "Disc %", "Promo EGP", "Save EGP", "Your 15%", "Rationale", "Status"]
    pick_widths  = [5, 42, 12, 8, 14, 12, 12, 50, 10]
    for j, (h, w) in enumerate(zip(pick_headers, pick_widths)):
        col_letter = chr(66 + j)
        ws4.column_dimensions[col_letter].width = w
        cell = ws4.cell(row=3, column=j+2)
        cell.value = h
        cell.font = Font(bold=True, color=WHITE, size=9)
        cell.fill = teal_fill()
        cell.alignment = center()
        cell.border = thin_border()
    ws4.row_dimensions[3].height = 18

    for i, o in enumerate(selected, start=4):
        disc  = _disc(o)
        promo = _promo(o)
        alc   = _alc(o)
        save  = alc - promo
        comm  = round(promo * commission, 2)
        fill  = lgreen_fill() if i % 2 == 0 else PatternFill("solid", fgColor="F0FAF4")
        vals  = [o.get("id", i-3), o.get("title",""), _party_size(o),
                 _disc_pct_str(o), promo, save, comm,
                 o.get("rationale",""), "Selected"]
        for j, v in enumerate(vals):
            cell = ws4.cell(row=i, column=j+2)
            cell.value = v
            cell.font = Font(size=9)
            cell.alignment = wrap()
            cell.border = thin_border()
            cell.fill = fill
        ws4.row_dimensions[i].height = 32

    # Backup separator
    sep_row = len(selected) + 4
    ws4.merge_cells(f"B{sep_row}:K{sep_row}")
    ws4[f"B{sep_row}"] = "BACKUPS — INTERNAL ONLY — DO NOT SHARE WITH PROVIDER"
    ws4[f"B{sep_row}"].font = Font(bold=True, color=WHITE)
    ws4[f"B{sep_row}"].fill = PatternFill("solid", fgColor="C0392B")
    ws4[f"B{sep_row}"].alignment = center()
    ws4.row_dimensions[sep_row].height = 18

    for i, o in enumerate(backups, start=sep_row+1):
        disc  = _disc(o)
        promo = _promo(o)
        alc   = _alc(o)
        save  = alc - promo
        comm  = round(promo * commission, 2)
        vals  = [f"B{i-sep_row}", o.get("title",""), _party_size(o),
                 _disc_pct_str(o), promo, save, comm,
                 o.get("rationale",""), "Backup"]
        for j, v in enumerate(vals):
            cell = ws4.cell(row=i, column=j+2)
            cell.value = v
            cell.font = Font(size=9, color="784212")
            cell.alignment = wrap()
            cell.border = thin_border()
            cell.fill = yellow_fill()
        ws4.row_dimensions[i].height = 32


def build_provider_xlsx(wb_out, data):
    """Build the provider-facing xlsx — selected offers only, max 15."""
    provider = data["provider"]
    commission = data.get("commission", 0.15)
    # Provider never sees backups — cap at 15, sorted ascending by promo price
    selected = [o for o in data.get("offers", []) if o.get("status", "Selected") == "Selected"][:15]
    selected = sorted(selected, key=lambda o: _promo(o))

    for sname in ["Sheet", "Sheet1"]:
        if sname in wb_out.sheetnames:
            del wb_out[sname]

    ws = wb_out.create_sheet("Provider Offers")

    # Header
    ws.merge_cells("A1:K1")
    ws["A1"] = f"{provider} x WeIN — Provider Negotiation Offers ({len(selected)} Selected Offers)"
    ws["A1"].font = Font(bold=True, size=14, color=WHITE)
    ws["A1"].fill = navy_fill()
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 30

    # KPI bar
    kpi_data = [
        ("Total Offers", str(len(selected))),
        ("Avg Discount", f"{sum(_disc(o) for o in selected)/len(selected):.0%}" if selected else "0%"),
        ("Highest Ticket", f"EGP {max((_alc(o) for o in selected), default=0):,.0f}"),
        ("Party Coverage", " · ".join(sorted(set(_party_size(o) for o in selected)))),
    ]
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 32
    ws.row_dimensions[4].height = 14

    kpi_cols = [1, 3, 5, 8]
    kpi_spans = [(1,2),(3,4),(5,7),(8,11)]
    for (start, end), (label, value) in zip(kpi_spans, kpi_data):
        sc = chr(64+start); ec = chr(64+end)
        if start != end:
            ws.merge_cells(f"{sc}3:{ec}3")
        ws[f"{sc}2"] = label
        ws[f"{sc}2"].font = Font(size=9, color="888888")
        ws[f"{sc}2"].alignment = center()
        ws[f"{sc}3"] = value
        ws[f"{sc}3"].font = Font(bold=True, size=12, color=NAVY)
        ws[f"{sc}3"].alignment = center()
        ws[f"{sc}3"].border = thin_border()

    # Column headers
    headers = ["#", "Offer Title", "Party Size", "Offer Contents", "Regular (EGP)",
               "Disc %", "Promo (EGP)", "Your 15%", "Guest Hook Line",
               "Validity / Terms"]
    widths  = [4, 40, 10, 55, 14, 8, 14, 12, 45, 40]
    col_letters = [chr(65+i) for i in range(len(headers))]
    for col, h, w in zip(col_letters, headers, widths):
        ws.column_dimensions[col].width = w
        cell = ws[f"{col}5"]
        cell.value = h
        cell.font = Font(bold=True, color=WHITE, size=9)
        cell.fill = blue_fill()
        cell.alignment = center()
        cell.border = thin_border()
    ws.row_dimensions[5].height = 18

    for i, o in enumerate(selected, start=6):
        disc  = _disc(o)
        promo = _promo(o)
        alc   = _alc(o)
        comm  = round(promo * commission, 2)
        fill  = lgrey_fill() if i % 2 == 0 else PatternFill("solid", fgColor="FAFAFA")

        vals = [o.get("id", i-5), o.get("title",""), _party_size(o),
                _contents(o), alc, _disc_pct_str(o), promo, comm,
                _hook_line(o), o.get("terms","")]

        for j, (col, v) in enumerate(zip(col_letters, vals)):
            cell = ws[f"{col}{i}"]
            if isinstance(v, list):
                cell.value = ', '.join(str(item) for item in v)
            else:
                cell.value = v
            cell.font = Font(size=9)
            cell.alignment = wrap()
            cell.border = thin_border()
            cell.fill = fill
        ws.row_dimensions[i].height = 50


def run(provider_name, vertical, output_dir, json_path, mode="full", explicit_version=None):
    """
    mode = 'full'       → build both xlsx files (new provider)
    mode = 'adjust'     → rebuild Filled Offer Template only, keep provider xlsx/PDFs unchanged
    mode = 'filled'     → rebuild Filled Offer Template only
    mode = 'provider'   → rebuild Provider Negotiation xlsx only
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)

    data["provider"] = provider_name
    data["vertical"] = vertical

    # Check minimum offer count for filled template
    all_offers = data.get("offers", [])
    selected = [o for o in all_offers if o.get("status", "Selected") == "Selected"]
    backups  = [o for o in all_offers if o.get("status", "Selected") == "Backup"]
    total    = len(selected) + len(backups)

    if total < 13 and mode in ("full", "adjust", "filled"):
        print(f"  WARNING: {total} offers found — minimum is 13 for Filled Offer Template.")
        print(f"  Add more offers to offer_data.json before rebuilding.")

    # ── File 1: Filled Offer Template (always rebuilt in full/adjust/filled mode) ──
    if mode in ("full", "adjust", "filled"):
        v1 = explicit_version if explicit_version is not None else next_version(output_dir, provider_name, "Filled Offer Template", "xlsx")
        wb1 = openpyxl.Workbook()
        build_filled_template(wb1, data)
        out1 = output_dir / f"{provider_name} - Filled Offer Template - Claude v{v1}.xlsx"
        wb1.save(out1)
        print(f"DONE: {out1.name}")

    # ── File 2: Provider Negotiation xlsx (only in full/provider mode — NOT in adjust) ──
    if mode in ("full", "provider"):
        v2 = explicit_version if explicit_version is not None else next_version(output_dir, provider_name, "Provider Negotiation Offers", "xlsx")
        wb2 = openpyxl.Workbook()
        build_provider_xlsx(wb2, data)
        out2 = output_dir / f"{provider_name} - Provider Negotiation Offers - Claude v{v2}.xlsx"
        wb2.save(out2)
        print(f"DONE: {out2.name}")

    if mode == "adjust":
        print(f"\nADJUST MODE: Filled Offer Template updated only.")
        print(f"Provider Negotiation xlsx and PDFs — NOT changed.")
    else:
        print(f"\nFiles saved to: {output_dir}")
        print("Run build_pdfs.py for PDF files 3 and 4.")


if __name__ == "__main__":
    if len(sys.argv) < 5:
        print("Usage: python build_offer_files.py <provider_name> <vertical> <output_dir> <json_path> [mode]")
        print("  mode: full (default) | adjust | filled | provider")
        sys.exit(1)
    mode = sys.argv[5] if len(sys.argv) > 5 else "full"
    explicit_version = int(sys.argv[6]) if len(sys.argv) > 6 else None
    run(sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4], mode, explicit_version=explicit_version)
